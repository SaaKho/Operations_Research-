import tkinter as tk
from tkinter import ttk
import pandas as pd
import numpy as np
from PIL import Image
from PIL import ImageTk


def SensitivityReports():
    WindowS = tk.Toplevel()
    canvas = tk.Canvas(WindowS, height=HEIGHT, width=WIDTH, background='lightblue')

    my_list = ["SensitivityReport1", "SensitivityReport2", "SensitivityReport3", "SensitivityReport4",
               "SensitivityReport5", "SensitivityReport6", "SensitivityReport7", "SensitivityReport8",
               "SensitivityReport9", "SensitivityReport10", "SensitivityReport11", "SensitivityReport12",
               "SensitivityReport13", "SensitivityReport14", "SensitivityReport15"]

    def check(*args):
        filename = options.get()
        print(f"the variable has changed to '{filename}'")

        from openpyxl import load_workbook
        wb = load_workbook(filename=filename + ".xlsx", read_only=True)
        ws = wb['Sheet1']  # connecting to work sheet

        l1 = ws.iter_rows(min_row=1, max_row=1, max_col=5, values_only=True)
        r_set = ws.iter_rows(min_row=2, max_row=5, values_only=True)

        l1 = [r for r in l1] 
        r_set = [r for r in r_set] 


        from tkinter import ttk  # for treeview
        # Using treeview widget
        trv = ttk.Treeview(WindowS, selectmode='browse')
        trv.grid(row=0, column=0, columnspan=3, padx=30, pady=20)

        trv['height'] = 30  # Number of rows to display, default is 10
        trv['show'] = 'headings'
        # column identifiers
        trv["columns"] = l1[0]
        # Defining headings, other option in tree
        # width of columns and alignment
        for i in l1[0]:
            trv.column(i, width=100, anchor='c')
        # Headings of respective columns
        for i in l1[0]:
            trv.heading(i, text=i)

        ## Adding data to treeview
        for dt in r_set:
            trv.insert("", 'end', iid=dt[0], values=dt)  # adding row

    options = tk.StringVar(WindowS)
    options.trace_add('write', check)
    options.set(my_list[0])  # default value


    om1 = tk.OptionMenu(WindowS, options, *my_list)
    om1.grid(row=5, column=2)

def AnswerReports():
    WindowA = tk.Toplevel()
    canvas = tk.Canvas(WindowA, height=HEIGHT, width=WIDTH, background='lightblue')

    my_list = ["AnswerReport1", "AnswerReport2", "AnswerReport3", "AnswerReport4",
               "AnswerReport5", "AnswerReport6", "AnswerReport7", "AnswerReport8",
               "AnswerReport9", "AnswerReport10", "AnswerReport11", "AnswerReport12",
               "AnswerReport13", "AnswerReport14", "AnswerReport15"]

    def check(*args):
        filename = options.get()
        print(f"the variable has changed to '{filename}'")

        from openpyxl import load_workbook
        wb = load_workbook(filename=filename + ".xlsx", read_only=True)
        ws = wb['Sheet1']  # connecting to work sheet

        l1 = ws.iter_rows(min_row=1, max_row=1, max_col=5, values_only=True)
        r_set = ws.iter_rows(min_row=2, max_row=5, values_only=True)

        l1 = [r for r in l1]  # Prepare list for column headers
        r_set = [r for r in r_set]  # Prepare list with data
        # wb.close()  # Close the workbook after reading
        # # print(l1) # to check the headers

        from tkinter import ttk  # for treeview
        # Using treeview widget
        trv = ttk.Treeview(WindowA, selectmode='browse')
        trv.grid(row=0, column=0, columnspan=3, padx=30, pady=20)

        trv['height'] = 30  # Number of rows to display, default is 10
        trv['show'] = 'headings'
        # column identifiers
        trv["columns"] = l1[0]
        # Defining headings, other option in tree
        # width of columns and alignment
        for i in l1[0]:
            trv.column(i, width=100, anchor='c')
        # Headings of respective columns
        for i in l1[0]:
            trv.heading(i, text=i)

        ## Adding data to treeview
        for dt in r_set:
            trv.insert("", 'end', iid=dt[0], values=dt)  # adding row

    options = tk.StringVar(WindowA)
    options.trace_add('write', check)
    options.set(my_list[0])  # default value

    om1 = tk.OptionMenu(WindowA, options, *my_list)
    om1.grid(row=5, column=2)


HEIGHT = 400
WIDTH = 500

ws = tk.Tk()
ws.title("Sadaan's Reports")
canvas = tk.Canvas(ws, height=HEIGHT, width=WIDTH, background='lightblue')
canvas.pack()

lbl = tk.Label(ws, text="Welcome to Saadan's OR Projects", fg='black', font=("Helvetica", 18), bg='lightblue')
lbl.place(x=60, y=100)


button1 = tk.Button(ws, text="Sensitivity Reports", bg='black', fg='White',
                   command=lambda: SensitivityReports())
button1.place(x=100, y=250)
button2 = tk.Button(ws, text="Answer Reports", bg='black', fg='White',
                   command=lambda: AnswerReports())
button2.place(x=300, y=250)

ws.mainloop()